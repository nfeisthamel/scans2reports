# Version Control
__version__ = "1.7.0"
def get_version() -> str:
    return __version__

import re
import pprint
import logging
import ipaddress
from PyQt5 import QtCore, QtGui, QtWidgets
from lxml import etree

""" Utilities module of static methods """
class Utils(object):
    """Utilities class """

    @staticmethod
    def update_status(application_path="", main_app=None, status=None, progress=None):
        if status:
            FORMAT = "[%(asctime)s ] %(levelname)s - %(filename)s; %(lineno)s: %(name)s.%(module)s.%(funcName)s(): %(message)s"
            logging.basicConfig(filename=f'{application_path}/scans2reports.log', level=logging.INFO, format=FORMAT)
            logging.info(status)        
            
        if main_app and main_app.main_window:
            if status:
                main_app.main_window.statusBar().showMessage(status)
            if progress is not None:
                main_app.main_window.progressBar.setValue( progress )
            QtGui.QGuiApplication.processEvents() 
                
    @staticmethod
    def fqdn(current_host):
        def _norm(s: str) -> str:
            s = (s or "").strip().strip('"').strip("'").rstrip('.').lower()
            # collapse internal whitespace
            return " ".join(s.split())

        def _non_ip_or_unknown(s: str) -> bool:
            s = (s or "").strip().lower()
            return bool(s) and s not in ("unknown", "n/a", "none") and not Utils.is_ip(s)

        fqdn_val = "UNKNOWN"

        # --- Primary: prefer Nessus HostProperties tags that commonly carry a name ---
        # Order chosen to prefer fully-qualified names when they exist.
        tag_candidates = [
            "./HostProperties/tag[@name='host-fqdn']/text()",
            "./HostProperties/tag[@name='host-rdns']/text()",     # reverse DNS name if present
            "./HostProperties/tag[@name='hostname']/text()",
            "./HostProperties/tag[@name='computer-name']/text()", # occasionally present
            "./HostProperties/tag[@name='netbios-name']/text()",  # Windows/NETBIOS
        ]
        for xp in tag_candidates:
            val = next(iter(current_host.xpath(xp)), "")
            val = _norm(str(val))
            if _non_ip_or_unknown(val):
                fqdn_val = val
                break

        # --- If we still only have IP/UNKNOWN, consider the host's @name attribute as a hint ---
        if not _non_ip_or_unknown(fqdn_val):
            name_attr = _norm(str(current_host.attrib.get("name", "")))
            if _non_ip_or_unknown(name_attr):
                fqdn_val = name_attr

        # --- If still empty/unknown/IP, try plugin 55472 "Device Hostname" ---
        if not _non_ip_or_unknown(fqdn_val):
            items_55472 = current_host.xpath(".//ReportItem[@pluginID='55472']")
            for ri in items_55472:
                output = (ri.findtext("plugin_output", default="") or "").strip()
                if not output:
                    continue

                lines = [ln.strip() for ln in output.splitlines()]

                # 1) Inline forms like "Hostname : foo", "Host name: foo", "FQDN: foo"
                for ln in lines:
                    # capture after colon; allow Hostname/Host name/FQDN/Fully Qualified Domain Name
                    m = re.match(r"(?i)^\s*(host\s*name|hostname|fqdn|fully\s+qualified.*domain\s+name)\s*:\s*(\S.*)$", ln)
                    if m:
                        cand = _norm(m.group(2))
                        if _non_ip_or_unknown(cand):
                            fqdn_val = cand
                            break
                if _non_ip_or_unknown(fqdn_val):
                    break

                # 2) Two-line forms: a label line followed by the value
                for i, ln in enumerate(lines):
                    if re.match(r"(?i)^\s*(host\s*name|hostname|fqdn|fully\s+qualified.*domain\s+name)\s*:?\s*$", ln):
                        if i + 1 < len(lines):
                            cand = _norm(lines[i + 1])
                            if _non_ip_or_unknown(cand):
                                fqdn_val = cand
                                break
                if _non_ip_or_unknown(fqdn_val):
                    break

        # --- Final normalization & log pairing ---
        if not _non_ip_or_unknown(fqdn_val):
            fqdn_val = "UNKNOWN"
            
        ip_val = next(iter(current_host.xpath("./HostProperties/tag[@name='host-ip']/text()")), 'UNKNOWN')

        # If we still couldn't resolve a hostname, fall back to the host IP
        if fqdn_val in ('', 'unknown', 'UNKNOWN'):
            fqdn_val = str(ip_val).lower()
         
        logging.info(f"[FQDN] Host resolved: fqdn='{fqdn_val}', ip='{ip_val}'")
         
        return fqdn_val

    @staticmethod
    def is_ip(val):
        try:
            ipaddress.ip_address(val)
            # logging.debug(f"[is_ip] Checked '{val}' => True")
            return True
        except ValueError:
            # logging.debug(f"[is_ip] Checked '{val}' => False")
            return False
    
    @staticmethod
    def clamp(number, minn, maxn):
        """ ensures integer stays within range """
        return max(min(maxn, number), minn)

    @staticmethod
    def parse_bool(source):
        """ parses non boolean values into boolean results """
        return bool(source.lower() in ['true', '1', 't', 'y', 'yes'])

    @staticmethod
    def status(source, form):
        """ will convert from multiple status formats into a specified format """
        if source is not None and form is not None:
            cross_walk_from = {
                'O'  : 0, 'FAIL':0, 'OPEN': 0, 'ONGOING' : 0,
                'NR' : 1, 'NOT_REVIEWED' : 1, 'NOT REVIEWED' : 1,
                'NA' : 2, 'NOTAPPLICABLE': 2, 'NOT_APPLICABLE': 2, 'NOT APPLICABLE' : 2,
                'C'  : 3, 'NOTAFINDING': 3, 'NOT_A_FINDING': 3, 'CLOSED' : 3, 'PASS' : 3, 'COMPLETED' : 3,
                'E'  : 4, 'ERROR' : 4
            }
            cross_walk_to = {
                'ABBREV' : {0: 'O', 1: 'NR', 2: 'NA', 3: 'C', 4: 'E'},
                'HUMAN' : {
                    0: 'Ongoing', 1: 'Not Reviewed', 2: 'Not Applicable', 3: 'Closed', 4: 'Error'
                },
                'RAW' : {
                    0: 'Open', 1: 'Not_Reviewed', 2: 'Not_Applicable', 3: 'Closed', 4: 'Error'
                },
            }
            return cross_walk_to[form][cross_walk_from[str(source).upper().strip()]]
        return 'Status Unknown'

    @staticmethod
    def risk_val(source, form):
        """ will convert from multiple risk formats into a specified format """
        if source is not None and form is not None:
            cross_walk_from = {
                '' : 0, 'UNKNOWN':0,
                'VL': 0, 'L': 1, 'M': 2, 'H': 3, 'VH': 4,
                'NONE': 0, 'INFO': 0, 'LOW': 1, 'MEDIUM': 2, 'HIGH': 3,
                'CRITICAL': 4, 'VERY HIGH' : 4,
                'CATIV': 0, 'CATIII': 1, 'CATII': 2, 'CATI': 3,
                'IV': 0, 'III': 1, 'II': 2, 'I': 3,
                'CAT IV': 0, 'CAT III': 1, 'CAT II': 2, 'CAT I': 3,
                'MODERATE': 2, 0: 0, 1: 1, 2: 2, 3: 3, 4: 4,
                '0':0, '1':1, '2':2, '3':3, '4':4
            }

            cross_walk_to = {
                'VL-VH': {
                    0: 'VL', 1: 'L', 2: 'M', 3: 'H', 4: 'VH',
                },
                'VL VH': {
                    0: 'Very Low', 1: 'Low', 2: 'Medium', 3: 'High', 4: 'Very High',
                },
                'POAM': {
                    0: 'Very Low', 1: 'Low', 2: 'Moderate', 3: 'High', 4: 'Very High',
                },
                'N-C': {
                    0: 'None', 1: 'Low', 2: 'Medium', 3: 'High', 4: 'Critical',
                },
                'CAT': {
                    0: 'CAT IV', 1: 'CAT III', 2: 'CAT II', 3: 'CAT I', 4: 'CAT I',
                },
                'MIN': {
                    0: 'IV', 1: 'III', 2: 'II', 3: 'I', 4: 'I',
                },
                'NUM':{
                    0: 0, 1: 1, 2: 2, 3: 3, 4: 4,
                },
            }
            
            # logging.debug(f"[risk_val] Raw input: {source}, Form: {form}")
            return cross_walk_to[form][cross_walk_from[str(source).upper().strip()]]
        return 'Risk Unknown'
        
    @staticmethod
    def write_mitigations_csv(report, output_dir):
        """
        Writes a 'mitigations.csv' file with 'Finding ID' and 'Mitigation' columns.
        Includes only rows with Status == 'Ongoing', or 'o', extracting the Security Checks field.
        """
        import csv
        import os

        output_path = os.path.join(output_dir, "mitigations.csv")

        try:
            with open(output_path, mode='w', newline='', encoding='utf-8') as f:
                writer = csv.writer(f)
                writer.writerow(['Finding ID', 'Mitigation'])  # Header row
                
                # Add example rows
                writer.writerow(['000000', 'This is an example, this works'])
                writer.writerow(['V-000000', 'So does this!'])

                for row in report:
                    status = row.get('Status', '').strip().lower()
                    if status in ['ongoing', 'o']:
                        raw_check = row.get('Security Checks', '').strip()
                        if not raw_check:
                            logging.info(f"[Utils-DEBUG] Skipping: Empty Security Checks for row with status '{status};")
                            continue

                        # Split into lines and apply selection logic
                        lines = [l.strip() for l in raw_check.splitlines() if l.strip()]
                        if len(lines) == 1:
                            finding_id = lines[0]
                        else:
                            v_lines = [line for line in lines if line.lower().startswith("V-")]
                            finding_id = v_lines[0] if v_lines else lines[-1]

                        if finding_id:
                            writer.writerow([finding_id, ''])

            logging.info(f"[Utils] Wrote mitigation template to {output_path}")
        except Exception as e:
            logging.error(f"[Utils] Failed to write mitigations.csv: {e}")
            
    @staticmethod
    def write_impacts_csv(report, output_dir):
        """
        Writes an 'impact.csv' file with 'Finding ID' and 'Impact Description' columns.
        Includes only rows with Status == 'Ongoing' or 'o', extracting the Security Checks field.
        Always prepends example entries.
        """
        import csv
        import os

        output_path = os.path.join(output_dir, "impacts.csv")

        try:
            with open(output_path, mode='w', newline='', encoding='utf-8') as f:
                writer = csv.writer(f, quoting=csv.QUOTE_MINIMAL)
                writer.writerow(['Finding ID', 'Impact Description'])  # Header row

                # Add example rows
                writer.writerow(['000000', 'This is an example, this works'])
                writer.writerow(['V-000000', 'So does this!'])

                for row in report:
                    status = row.get('Status', '').strip().lower()
                    if status in ['ongoing', 'o']:
                        raw_check = row.get('Security Checks', '').strip()
                        if not raw_check:
                            logging.info(f"[Utils-DEBUG] Skipping: Empty Security Checks for row with status '{status}'")
                            continue

                        lines = [l.strip() for l in raw_check.splitlines() if l.strip()]
                        if len(lines) == 1:
                            finding_id = lines[0]
                        else:
                            v_lines = [line for line in lines if line.lower().startswith("v-")]
                            finding_id = v_lines[0] if v_lines else lines[-1]

                        if finding_id:
                            writer.writerow([finding_id, ''])  # Blank Impact Description field

            logging.info(f"[Utils] Wrote impact template to {output_path}")
        except Exception as e:
            logging.error(f"[Utils] Failed to write impact.csv: {e}")
            
    @staticmethod
    def write_resources_csv(report, output_dir):
        """
        Writes an 'resource.csv' file with 'Finding ID' and 'Resource Description' columns.
        Includes only rows with Status == 'Ongoing' or 'o', extracting the Security Checks field.
        Always prepends example entries.
        """
        import csv
        import os

        output_path = os.path.join(output_dir, "resources.csv")

        try:
            with open(output_path, mode='w', newline='', encoding='utf-8') as f:
                writer = csv.writer(f, quoting=csv.QUOTE_MINIMAL)
                writer.writerow(['Finding ID', 'Resources Required'])  # Header row

                # Add example rows
                writer.writerow(['000000', 'This is an example, this works'])
                writer.writerow(['V-000000', 'So does this!'])

                for row in report:
                    status = row.get('Status', '').strip().lower()
                    if status in ['ongoing', 'o']:
                        raw_check = row.get('Security Checks', '').strip()
                        if not raw_check:
                            logging.info(f"[Utils-DEBUG] Skipping: Empty Security Checks for row with status '{status}'")
                            continue

                        lines = [l.strip() for l in raw_check.splitlines() if l.strip()]
                        if len(lines) == 1:
                            finding_id = lines[0]
                        else:
                            v_lines = [line for line in lines if line.lower().startswith("v-")]
                            finding_id = v_lines[0] if v_lines else lines[-1]

                        if finding_id:
                            writer.writerow([finding_id, ''])  # Blank Resource Required field

            logging.info(f"[Utils] Wrote resource template to {output_path}")
        except Exception as e:
            logging.error(f"[Utils] Failed to write resource.csv: {e}")
            
    @staticmethod
    def enrich_hwsw_workbook(prior_path, new_path, output_path=None):
        import openpyxl
        import logging
        import re
        
        from copy import copy
        from openpyxl.worksheet.cell_range import CellRange, MultiCellRange

        def get_str(val):
            return str(val).strip() if val is not None else ''

        def normalize_header(header):
            try:
                s = str(header).strip() if header is not None else ''
            except Exception:
                s = ''
            if not s:
                return ''
            return re.sub(r'[\W_]+', '', s.lower())

        header_map = {
            "machinenamerequired": "hostname",
            "assetname": "hostname",
            "criticalinformationsystemasset": "criticalinformationsystem",
            "osiostfwversion": "osversion",
            "assetipaddress": "ipaddress",
            "macaddress": "macaddress",
            "componenttype": "componenttype",
            "hosts": "hostname",
            "parentsystem": "hostname",
            "name": "softwarename",
            "version": "version"
        }

        def resolve_header(header):
            norm = normalize_header(header)
            return header_map.get(norm, norm)

        def find_header_row(ws, expected_headers=None):
            if expected_headers is None:
                expected_headers = {"ipaddress", "hostname"}

            for i, row in enumerate(ws.iter_rows(min_row=0, max_row=10, values_only=False)):
                headers = [normalize_header(cell.value) for cell in row]
                if any(h in headers for h in expected_headers):
                    return i + 1, [get_str(cell.value) for cell in row]
            for i, row in enumerate(ws.iter_rows(min_row=6, max_row=7, values_only=False)):
                if str(row[0].value).strip() == "#":
                    return i + 1, [get_str(cell.value) for cell in row]
            raise Exception("Could not detect header row.")

        def normalize_key(key):
            return tuple(get_str(part).lower() for part in key)

        logging.info("Enriching HWSW workbook: %s with prior: %s", new_path, prior_path)

        prior_wb = openpyxl.load_workbook(prior_path, data_only=False)
        new_wb = openpyxl.load_workbook(new_path)

        if "Hardware" not in prior_wb.sheetnames or "Hardware" not in new_wb.sheetnames:
            logging.warning("Missing 'Hardware' sheet in one of the workbooks.")
            return

        prior_ws = prior_wb["Hardware"]
        new_ws = new_wb["Hardware"]

        prior_header_row, prior_headers = find_header_row(prior_ws)
        new_header_row, new_headers = find_header_row(new_ws)

        norm_prior_headers = {resolve_header(h): i for i, h in enumerate(prior_headers) if h}
        norm_new_headers = {resolve_header(h): i for i, h in enumerate(new_headers) if h}

        logging.info(f"[headers] new_headers: {new_headers}")
        logging.info(f"[headers] resolved new: {norm_new_headers}")
        logging.info(f"[headers] prior_headers: {prior_headers}")
        logging.info(f"[headers] resolved prior: {norm_prior_headers}")

        # Step 1: Build match dictionaries
        prior_soft_match = {}
        prior_hard_match = {}

        for row in prior_ws.iter_rows(min_row=prior_header_row + 1, values_only=False):
            ip_idx = norm_prior_headers.get("ipaddress", -1)
            mac_idx = norm_prior_headers.get("macaddress", -1)
            ip = get_str(row[ip_idx].value) if 0 <= ip_idx < len(row) else ''
            mac = get_str(row[mac_idx].value) if 0 <= mac_idx < len(row) else ''
            if not ip:
                continue
            prior_soft_match[normalize_key((ip,))] = row
            if mac:
                prior_hard_match[normalize_key((ip, mac))] = row

        enriched_rows = 0
        appended_rows = 0
        asset_name_updated = 0
        matched_keys = set()
        written_keys = set()
        log_lines = []

        # Phase 1: Enrich rows in-place
        for row in new_ws.iter_rows(min_row=new_header_row + 1):
            ip = get_str(row[norm_new_headers.get("ipaddress")].value) if "ipaddress" in norm_new_headers else ''
            mac = get_str(row[norm_new_headers.get("macaddress")].value) if "macaddress" in norm_new_headers else ''
            key_hard = normalize_key((ip, mac))
            key_soft = normalize_key((ip,))
            matched_row = None
            match_type = None
            row_updated = False

            if key_hard in prior_hard_match:
                matched_row = prior_hard_match[key_hard]
                match_type = "hard"
            elif key_soft in prior_soft_match:
                matched_row = prior_soft_match[key_soft]
                match_type = "soft"

            key_desc = f"{ip} / {mac}"
            norm_written_key = normalize_key((ip, mac))

            if norm_written_key in written_keys:
                logging.debug(f"[phase1 skip] already written: {norm_written_key}")
                continue

            if matched_row:
                logging.info(f"[match] {match_type} match found for {key_desc}")
                for hname, col_idx in norm_new_headers.items():
                    if col_idx == 0:
                        continue
                    prior_idx = norm_prior_headers.get(hname, -1)
                    if prior_idx < 0 or prior_idx >= len(matched_row):
                        continue

                    cell = row[col_idx]
                    new_val = matched_row[prior_idx].value

                    if hname == "hostname":
                        ip_idx_new = norm_new_headers.get("ipaddress")
                        ip_idx_old = norm_prior_headers.get("ipaddress")
                        name_idx_old = norm_prior_headers.get("hostname")
                        if ip_idx_new is not None and ip_idx_old is not None and name_idx_old is not None:
                            ip_new = get_str(row[ip_idx_new].value)
                            ip_old = get_str(matched_row[ip_idx_old].value)
                            old_name = get_str(matched_row[name_idx_old].value)
                            current_val = get_str(cell.value)
                            logging.debug(f"[hostname check] ip_new='{ip_new}', ip_old='{ip_old}', old_name='{old_name}', current_val='{current_val}'")

                            if ip_new == ip_old and old_name and (current_val == ip_new or current_val == ''):
                                cell.value = old_name
                                row_updated = True
                                asset_name_updated += 1
                                log_lines.append(f"[hostname fix] {key_desc} - '{current_val}' -> '{old_name}'")
                            else:
                                logging.debug(f"[hostname skip] ip_match={ip_new == ip_old}, old_name_ok={bool(old_name)}, cur_is_ip_or_blank={current_val == ip_new or current_val == ''}")
                        continue

                    if get_str(cell.value) in ['', 'N/A'] and new_val not in [None, '', 'N/A']:
                        cell.value = new_val
                        row_updated = True
                        log_lines.append(f"[{match_type} fill] {key_desc} -> {hname} = '{new_val}'")

            if matched_row:
                matched_keys.add(key_hard)
                matched_keys.add(key_soft)
                written_keys.add(norm_written_key)
            
            if row_updated:
                enriched_rows += 1
            else:
                logging.info(f"[no match] {key_desc}")

        # Phase 2: Append unmatched rows
        appended_keys = set()
        for match_dict, label in [(prior_hard_match, "hard"), (prior_soft_match, "soft")]:
            for key, row in match_dict.items():
                row_vals = [''] * new_ws.max_column
                for hname, new_col_idx in norm_new_headers.items():
                    if hname in norm_prior_headers:
                        prior_idx = norm_prior_headers[hname]
                        if prior_idx < len(row):
                            row_vals[new_col_idx] = row[prior_idx].value

                # Safely extract string values for header check
                header_hits = 0
                for cell in row_vals:
                    val = get_str(cell)
                    if val and normalize_header(val) in norm_new_headers:
                        header_hits += 1

                if header_hits >= 3:
                    logging.debug(f"[append skip] suspected header row (hits={header_hits}): {row_vals}")
                    continue
                    
                norm_key = normalize_key(key)
                ip_idx = norm_prior_headers.get("ipaddress", -1)
                mac_idx = norm_prior_headers.get("macaddress", -1)
                ip = get_str(row[ip_idx].value) if 0 <= ip_idx < len(row) else ''
                mac = get_str(row[mac_idx].value) if 0 <= mac_idx < len(row) else ''
                ip_only_key = normalize_key((get_str(row[ip_idx].value),))
                norm_written_key = normalize_key((ip, mac))

                if norm_written_key in written_keys or norm_key in appended_keys or norm_key in matched_keys or ip_only_key in matched_keys:
                    logging.debug(f"[append skip] already written: {norm_key}")
                    continue

                row_vals[0] = ''
                new_ws.append(row_vals)

                new_row_idx = new_ws.max_row
                ref_row_idx = new_row_idx - 1
                
                for col_idx in range(1, new_ws.max_column + 1):
                    new_cell = new_ws.cell(row=new_row_idx, column=col_idx)
                    ref_cell = new_ws.cell(row=ref_row_idx, column=col_idx)
                    try:
                        new_cell.font = copy(ref_cell.font)
                        new_cell.border = copy(ref_cell.border)
                        new_cell.fill = copy(ref_cell.fill)
                        new_cell.number_format = ref_cell.number_format
                        new_cell.protection = copy(ref_cell.protection)
                        new_cell.alignment = copy(ref_cell.alignment)
                    except Exception as e:
                        logging.warning(f"[formatting error] row={new_row_idx}, col={col_idx} — {e}")
                        
                # Copy data validation from ref_row_idx to new_row_idx
                for dv in list(new_ws.data_validations.dataValidation):
                    for cell_range in dv.ranges:
                        for col in range(1, new_ws.max_column + 1):
                            col_letter = openpyxl.utils.get_column_letter(col)
                            ref_cell = f"{col_letter}{ref_row_idx}"
                            new_cell = f"{col_letter}{new_row_idx}"
                            if ref_cell in cell_range:
                                # Safely create a new MultiCellRange combining the existing one with the new cell
                                if not any(new_cell in cr for cr in dv.ranges.ranges):
                                    dv.ranges = MultiCellRange(set(dv.ranges.ranges).union({CellRange(new_cell)}))

                appended_rows += 1
                appended_keys.add(norm_key)
                written_keys.add(norm_written_key)
                log_lines.append(f"[append] unmatched {label} row appended for key={key}")
                logging.info(f"[ENRICHMENT-STATS] enriched={enriched_rows}, hostnames_updated={asset_name_updated}, appended={appended_rows}")
                
                
        # --- Phase 3: Enrich Software tab with missing prior host/software rows ---
        if "Software" in prior_wb.sheetnames and "Software" in new_wb.sheetnames:
            prior_ws_software = prior_wb["Software"]
            new_ws_software = new_wb["Software"]

            # Detect headers in both
            expected_software_headers = {"softwarename", "version", "hostname"}

            prior_header_row, prior_headers = find_header_row(
                prior_ws_software, expected_headers=expected_software_headers
            )
            new_header_row, new_headers = find_header_row(
                new_ws_software, expected_headers=expected_software_headers
            )

            prior_headers_map = {resolve_header(h): i for i, h in enumerate(prior_headers)}
            new_headers_map = {resolve_header(h): i for i, h in enumerate(new_headers)}

            def extract_triplets(ws, header_map, header_row):
                triplets = set()
                name_idx = header_map.get("softwarename")
                version_idx = header_map.get("version")
                host_idx = header_map.get("hostname")

                if None in (name_idx, version_idx, host_idx):
                    logging.error(f"[SOFTWARE] Missing expected columns in header_map: {header_map}")
                    return triplets  # Early exit

                for row in ws.iter_rows(min_row=header_row + 1, values_only=True):
                    name = get_str(row[name_idx]) if name_idx < len(row) else ''
                    version = get_str(row[version_idx]) if version_idx < len(row) else ''
                    hosts = get_str(row[host_idx]) if host_idx < len(row) else ''
                    for host in hosts.split(","):
                        host = host.strip()
                        if host:
                            triplets.add((name, version, host))
                return triplets

            prior_software = extract_triplets(prior_ws_software, prior_headers_map, prior_header_row)
            new_software = extract_triplets(new_ws_software, new_headers_map, new_header_row)

            missing_entries = sorted(prior_software - new_software)
            appended_software = 0

            if missing_entries:
                logging.info(f"[SOFTWARE] Found {len(missing_entries)} missing software entries to append.")

                for name, version, host in missing_entries:
                    row_vals = [''] * new_ws_software.max_column
                    for hname, col_idx in new_headers_map.items():
                        if hname == "softwarename":
                            row_vals[col_idx] = name
                        elif hname == "version":
                            row_vals[col_idx] = version
                        elif hname == "hostname":
                            row_vals[col_idx] = host

                    new_ws_software.append(row_vals)
                    new_row_idx = new_ws_software.max_row
                    ref_row_idx = new_row_idx - 1

                    # Copy formatting from previous row
                    for col_idx in range(1, new_ws_software.max_column + 1):
                        new_cell = new_ws_software.cell(row=new_row_idx, column=col_idx)
                        ref_cell = new_ws_software.cell(row=ref_row_idx, column=col_idx)
                        try:
                            new_cell.font = copy(ref_cell.font)
                            new_cell.border = copy(ref_cell.border)
                            new_cell.fill = copy(ref_cell.fill)
                            new_cell.number_format = ref_cell.number_format
                            new_cell.protection = copy(ref_cell.protection)
                            new_cell.alignment = copy(ref_cell.alignment)
                        except Exception as e:
                            logging.warning(f"[SW FORMAT] row={new_row_idx}, col={col_idx} — {e}")

                    # Copy data validation if applicable
                    for dv in list(new_ws_software.data_validations.dataValidation):
                        cells_to_add = []
                        for cell_range in dv.ranges:
                            for col in range(1, new_ws_software.max_column + 1):
                                col_letter = openpyxl.utils.get_column_letter(col)
                                ref_cell = f"{col_letter}{ref_row_idx}"
                                new_cell = f"{col_letter}{new_row_idx}"
                                if ref_cell in cell_range and new_cell not in cell_range:
                                    cells_to_add.append(CellRange(new_cell))
                        for cr in cells_to_add:
                            dv.ranges.add(cr)

                    appended_software += 1
                    logging.debug(f"[SOFTWARE append] name='{name}', version='{version}', host='{host}'")

                log_lines.append(f"[SOFTWARE] {appended_software} legacy software rows appended.")
                logging.info(f"[SOFTWARE-STATS] appended={appended_software}")
            else:
                logging.info("[SOFTWARE] No legacy software entries to append.")
        else:
            logging.warning("[SOFTWARE] One or both Software tabs not found. Skipping software enrichment.")


        save_path = output_path or new_path
        new_wb.save(save_path)
        logging.info(f"[ENRICHMENT-SAVE] Workbook Saved to: {save_path}")
